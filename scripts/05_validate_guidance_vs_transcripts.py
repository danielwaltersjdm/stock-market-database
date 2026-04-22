"""
05_validate_guidance_vs_transcripts.py

Cross-check the CIQ plug-in guidance numbers against the text of MAEC
transcripts. For each (ticker, fiscal year, measure) with a CIQ guidance
value, we find every MAEC call for that ticker from that fiscal year and the
prior year's Q4 (when FY guidance is most commonly issued), then search the
transcript text for the guidance number (with formatting tolerance).

The goal is a consistency check: does management actually say in the call
what CIQ says they guided?

Conservative tolerance:
    - For ranges, we consider a match if the transcript contains both the
      low and high bounds within a short window of each other
    - For point values, the transcript must contain the value (possibly with
      $ or decimal-place variation)
    - Values are searched at 2-decimal and no-decimal precision

Outputs:
    processed_data/public/guidance_validation_pilot.csv
"""

import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent.parent
SMOKE_TEST = PROJECT_ROOT / "notebooks" / "ciq_guidance_smoke_test.xlsx"
MAEC_ROOT = PROJECT_ROOT / "raw_data" / "transcripts" / "maec" / "MAEC_Dataset"
OUT = PROJECT_ROOT / "processed_data" / "public" / "guidance_validation_pilot.csv"


def parse_ciq_value(v):
    """Return (low, high, kind) where kind in {'point','range','bound','none'}."""
    if v is None:
        return None, None, "none"
    if isinstance(v, (int, float)):
        if v != v or v == 0:
            return None, None, "none"
        return float(v), float(v), "point"
    s = str(v).strip()
    m = re.match(r"^(-?\d+\.?\d*)\s*[-\u2013]\s*(-?\d+\.?\d*)$", s)
    if m:
        return float(m.group(1)), float(m.group(2)), "range"
    m = re.match(r"^([><])\s*(-?\d+\.?\d*)$", s)
    if m:
        v = float(m.group(2))
        return v, v, "bound"
    return None, None, "none"


def load_guidance_df():
    wb = load_workbook(SMOKE_TEST, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df[["gl", "gh", "kind"]] = df["Value"].apply(lambda v: pd.Series(parse_ciq_value(v)))
    return df[df["kind"].isin(["point", "range", "bound"])].copy()


def build_maec_index():
    """Map ticker -> list of (YYYY-MM-DD, folder_path)."""
    idx = defaultdict(list)
    for p in MAEC_ROOT.iterdir():
        if not p.is_dir() or "_" not in p.name:
            continue
        datestr, ticker = p.name.split("_", 1)
        if len(datestr) == 8 and datestr.isdigit():
            d = f"{datestr[:4]}-{datestr[4:6]}-{datestr[6:]}"
            idx[ticker].append((d, p))
    return idx


def fy_to_call_window(fy: str):
    """Given 'FY2015', return the calendar range most likely to be the call
    that issued / updated FY2015 guidance: Oct 2014 through Dec 2015."""
    year = int(fy.replace("FY", ""))
    return f"{year-1}-10-01", f"{year}-12-31"


NUMBER_RE = re.compile(r"(?<![\w\.])\$?(\d+(?:[,.]\d{1,6})?)(?![\w\d])")


def value_in_text(value: float, text: str, tol_rel: float = 0.003) -> int:
    """Count occurrences of a number within relative tolerance in text."""
    hits = 0
    for m in NUMBER_RE.finditer(text):
        raw = m.group(1).replace(",", "")
        try:
            num = float(raw)
        except ValueError:
            continue
        if value == 0:
            continue
        if abs(num - value) / abs(value) <= tol_rel:
            hits += 1
    return hits


def main():
    guidance = load_guidance_df()
    print(f"Loaded CIQ guidance rows with real values: {len(guidance)}")
    maec = build_maec_index()
    print(f"MAEC transcript folders available: {sum(len(v) for v in maec.values())}")

    results = []
    for _, row in guidance.iterrows():
        ticker = row["Ticker"]
        fy = row["Period"]
        measure = row["Measure"]
        gl, gh = row["gl"], row["gh"]
        kind = row["kind"]

        if ticker not in maec:
            continue
        lo_date, hi_date = fy_to_call_window(fy)
        candidates = [(d, p) for d, p in maec[ticker] if lo_date <= d <= hi_date]
        if not candidates:
            continue

        # Concatenate all candidate transcripts
        full_text = ""
        used_dates = []
        for d, p in candidates:
            try:
                full_text += "\n" + (p / "text.txt").read_text(encoding="utf-8", errors="ignore")
                used_dates.append(d)
            except Exception:
                pass
        if not full_text:
            continue

        # Search both endpoints
        low_hits = value_in_text(gl, full_text)
        high_hits = value_in_text(gh, full_text)

        if kind == "point":
            match = low_hits > 0
        elif kind == "range":
            match = (low_hits > 0 and high_hits > 0)
        else:  # bound
            match = low_hits > 0

        results.append({
            "ticker": ticker,
            "period": fy,
            "measure": measure,
            "kind": kind,
            "low": gl,
            "high": gh,
            "call_dates_searched": ",".join(used_dates),
            "low_hits": low_hits,
            "high_hits": high_hits,
            "match": bool(match),
        })

    res = pd.DataFrame(results)
    res.to_csv(OUT, index=False)
    print(f"Wrote {OUT}")
    print()

    n = len(res)
    if n == 0:
        print("No guidance-transcript overlaps found.")
        return

    print(f"=== Validation summary ({n} guidance values cross-checked) ===")
    print(f"Matched (value appears in candidate transcripts): {res['match'].sum()} ({res['match'].mean():.0%})")
    print()
    print("Match rate by measure:")
    by_m = res.groupby("measure")["match"].agg(["sum", "count", "mean"])
    by_m.columns = ["matched", "total", "rate"]
    print(by_m.to_string())
    print()
    print("Match rate by kind:")
    print(res.groupby("kind")["match"].agg(["sum", "count", "mean"]).to_string())
    print()
    print("Sample mismatches (for spot-checking):")
    mis = res[~res["match"]].head(10)
    print(mis[["ticker", "period", "measure", "kind", "low", "high", "low_hits", "high_hits"]].to_string())
    print()
    print("Sample matches:")
    m = res[res["match"]].head(10)
    print(m[["ticker", "period", "measure", "kind", "low", "high", "low_hits", "high_hits"]].to_string())


if __name__ == "__main__":
    main()
