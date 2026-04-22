"""
06_run_ciq_chunks.py

Drive Excel via xlwings to refresh the CIQ plug-in on each chunk workbook,
save, and read the results without user intervention.

Usage
-----
    # Run all chunks that haven't been refreshed yet (based on a completion marker)
    python scripts/06_run_ciq_chunks.py

    # Run specific chunk(s)
    python scripts/06_run_ciq_chunks.py --chunks 2 3 4

    # Force re-refresh a chunk that was already done
    python scripts/06_run_ciq_chunks.py --chunks 1 --force

Assumes:
    - Excel installed
    - CIQ plug-in installed and signed in
    - The chunk workbooks live at notebooks/ciq_guidance_chunk_NN.xlsx
"""

import argparse
import re
import sys
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent.parent
NOTEBOOKS = PROJECT_ROOT / "notebooks"
PUBLIC_DIR = PROJECT_ROOT / "processed_data" / "public"
PUBLIC_DIR.mkdir(parents=True, exist_ok=True)

REFRESH_WAIT_SECONDS = 10   # initial wait after triggering Refresh
MAX_POLL_SECONDS = 300      # hard cap — 5 minutes per chunk
POLL_INTERVAL = 5


def parse_value(v):
    """Categorize and parse one cell value to match downstream schema."""
    if v is None:
        return None, "none"
    if isinstance(v, (int, float)):
        if v != v:
            return None, "nan"
        if v == 0:
            return None, "zero_as_na"
        return float(v), "number"
    s = str(v).strip()
    if s == "":
        return None, "empty"
    if "invalid" in s.lower():
        return None, "invalid"
    s2 = re.sub(r"\((-?\d+\.?\d*)\)", r"-\1", s)  # handle (0.53) negatives
    m = re.match(r"^(-?\d+\.?\d*)\s*[-\u2013]\s*(-?\d+\.?\d*)$", s2)
    if m:
        return (float(m.group(1)) + float(m.group(2))) / 2, "range"
    m = re.match(r"^[><]\s*(-?\d+\.?\d*)$", s2)
    if m:
        return float(m.group(1)), "bound"
    return None, "other"


def is_loading(value) -> bool:
    """CIQ plug-in typically writes 'Loading...' or '#BUSY' while async refresh runs."""
    if value is None:
        return False
    s = str(value).strip().lower()
    return "loading" in s or "busy" in s or s == "#n/a requesting data..."


def refresh_and_save_with_xlwings(xlsx_path: Path) -> dict:
    """Open, refresh all, wait for completion, save, close. Return timing info."""
    import xlwings as xw

    t0 = time.time()
    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    app.screen_updating = True  # we want the plug-in to paint

    try:
        wb = app.books.open(str(xlsx_path))
        ws = wb.sheets[0]

        # Trigger full recalculation (CalculateFullRebuild ~ Ctrl+Alt+F9)
        app.api.CalculateFullRebuild()
        time.sleep(REFRESH_WAIT_SECONDS)

        # Poll column E for lingering "Loading..." or #BUSY values
        start = time.time()
        last_row = ws.used_range.last_cell.row
        col_range = f"E2:E{last_row}"
        while time.time() - start < MAX_POLL_SECONDS:
            vals = ws.range(col_range).value
            if vals is None:
                vals = []
            elif not isinstance(vals, list):
                vals = [vals]
            # flatten if list-of-lists
            if vals and isinstance(vals[0], list):
                vals = [v for row_ in vals for v in row_]
            n_loading = sum(1 for v in vals if is_loading(v))
            if n_loading == 0:
                break
            print(f"  waiting for {n_loading} cells to finish loading...", flush=True)
            time.sleep(POLL_INTERVAL)

        # Save in place (xlsx preserves values written by plug-in)
        wb.save()
        wb.close()
    finally:
        app.quit()

    return {"elapsed_seconds": round(time.time() - t0, 1)}


def parse_chunk_file(xlsx_path: Path) -> pd.DataFrame:
    """Re-read the saved file and produce a parsed CSV."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df[["parsed_value", "category"]] = df["Value"].apply(
        lambda v: pd.Series(parse_value(v))
    )
    df["Value"] = df["Value"].astype(str)
    return df


def summary(df: pd.DataFrame, label: str = "") -> None:
    real_cats = {"number", "range", "bound"}
    real = df[df["category"].isin(real_cats)]
    firms_total = df["Ticker"].nunique()
    firms_with = real["Ticker"].nunique()
    print(f"  {label}: {len(real):,}/{len(df):,} real ({100*len(real)/len(df):.0f}%); "
          f"{firms_with}/{firms_total} firms")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--chunks", nargs="*", type=int, default=None,
                    help="Specific chunk numbers to run (default: all that lack a parsed CSV)")
    ap.add_argument("--force", action="store_true",
                    help="Re-run even if a parsed CSV already exists")
    args = ap.parse_args()

    all_chunks = sorted(NOTEBOOKS.glob("ciq_guidance_chunk_*.xlsx"))
    if not all_chunks:
        sys.exit("No chunk workbooks found in notebooks/")

    # Determine which to run
    def chunk_num(p):
        m = re.search(r"chunk_(\d+)", p.name)
        return int(m.group(1)) if m else -1

    if args.chunks:
        targets = [p for p in all_chunks if chunk_num(p) in set(args.chunks)]
    else:
        targets = []
        for p in all_chunks:
            out_csv = PUBLIC_DIR / f"ciq_guidance_chunk_{chunk_num(p):02d}.csv"
            if out_csv.exists() and not args.force:
                continue
            targets.append(p)

    if not targets:
        print("Nothing to run (all chunks have parsed CSVs already; pass --force to redo).")
        return

    print(f"Will run {len(targets)} chunk(s): {[p.name for p in targets]}")

    for p in targets:
        n = chunk_num(p)
        print(f"\n=== Chunk {n:02d}: {p.name} ===", flush=True)
        try:
            info = refresh_and_save_with_xlwings(p)
            print(f"  refresh+save OK in {info['elapsed_seconds']}s", flush=True)
        except Exception as e:
            print(f"  refresh FAILED: {type(e).__name__}: {e}", flush=True)
            continue

        df = parse_chunk_file(p)
        out_csv = PUBLIC_DIR / f"ciq_guidance_chunk_{n:02d}.csv"
        df.to_csv(out_csv, index=False)
        print(f"  saved {out_csv}")
        summary(df, f"chunk {n:02d}")

    # Aggregate all existing parsed chunks
    csvs = sorted(PUBLIC_DIR.glob("ciq_guidance_chunk_*.csv"))
    if csvs:
        print("\n=== Aggregate across all parsed chunks ===")
        all_df = pd.concat([pd.read_csv(c) for c in csvs], ignore_index=True)
        all_out = PUBLIC_DIR / "ciq_guidance_all.csv"
        all_df.to_csv(all_out, index=False)
        summary(all_df, "AGGREGATE")
        print(f"  wrote {all_out}  ({len(all_df):,} rows)")


if __name__ == "__main__":
    main()
